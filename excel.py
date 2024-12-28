from coredb import mod_sql_products, check_prod, mod_sql_order, mod_sql_payment, mod_sql_item_orders, mod_sql_service\
    , mod_sql_service_pay, mod_sql_reimburs, mod_sql_contractor, mod_sql_contractor_pay, mod_sql_rent
import time
import openpyxl
import datetime
import copy
flgDebug = False;
class CoreXLSX:
    def migration_product_in_postgres(self):
        wb = openpyxl.load_workbook("Code_Product_SQL.xlsx")
        ws = wb['Migration_to_psql']
        start = time.time()
        rows = ws.iter_rows(min_row=11408, max_row=11441, min_col=2, max_col=20)
        for c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14, c15, c16, c17, c18, c19 in rows:
            d = {'id_product': c1.value, 'type': c16.value, 'c1': c17.value, 'c2': c18.value, 'c3': c19.value,
                 'unit': c10.value, 'ratio': c11.value, 'product': c2.value, 'provider_id': c3.value}
            data = mod_sql_products(d);
        stop = time.time()
        s = stop - start
        print('\n', round(s, 2), 'сек.')
    def read_xlsx_products(file_name):
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        max_col = ws.max_column
        str_pet = ws['f2']
        data = []
        if check_petrovich(str_pet) == "ok":
            rows = ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=3, max_col=4)
            data = check_prod(rows)
        else:
            if max_col > 2:
                return {'status': 'err', 'data': []}
            else:
                rows = ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2)
                if rows:
                    data = check_prod(rows)
        return {'status': 'ok', 'data': data}
    def migration_order_in_postgres(self):
        wb = openpyxl.load_workbook("C:\\Users\\Master\\Desktop\\РАБОТА\\Transfer_to_Postgres\\trans_sql.xlsx")
        ws1 = wb['sett']
        id_contract = copy.copy(ws1['a2'].value)
        id_refs = copy.copy(ws1['b2'].value)
        print(id_contract)
        ws2 = wb['Group_SQL']
        print(ws2.max_row)
        dict = []
        rows = ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=8)
        for c1, c2, c3, c4, c5, c6, c7, c8 in rows:
            d = {'id_name': c1.value, 'order_date': c2.value, 'id_provider': c4.value,
                'order_total': round(c5.value, 2), 'id_contract': id_contract,
                'created': datetime.datetime.now(), 'update': datetime.datetime.now(), 'type': 'CONTR',
                 'id_adress_refreshment': id_refs}
        #     # d = {'id_name': c1.value, 'order_date': c2.value}
            dict.append(d)
        print(dict)
        data = mod_sql_order(dict);


    def migration_payment_in_postgres(self):
        wb = openpyxl.load_workbook("C:\\Users\\Master\\Desktop\\РАБОТА\\Transfer_to_Postgres\\trans_sql.xlsx")
        ws1 = wb['sett']
        id_contract = copy.copy(ws1['a2'].value)
        id_refs = copy.copy(ws1['b2'].value)
        ws2 = wb['Group_SQL']
        print(ws2.max_row)
        dict = []
        rows = ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=11)
        for c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11 in rows:
            d = {'id_name': c1.value, 'id_contract': id_contract, 'payment_check': round(c5.value, 2),
                 'payment_date': c2.value, 'account_number': c8.value, 'id_company': c7.value,
                 'created': datetime.datetime.now(), 'update': datetime.datetime.now(), 'id_refs': id_refs}
            dict.append(d)
        print(dict)
        data = mod_sql_payment(dict);


    def migration_items_in_postgres(self):
        wb = openpyxl.load_workbook("C:\\Users\\Master\\Desktop\\РАБОТА\\Transfer_to_Postgres\\trans_sql.xlsx")
        ws1 = wb['sett']
        id_contract = copy.copy(ws1['a2'].value)
        id_refs = copy.copy(ws1['b2'].value)
        ws2 = wb['items']
        print(ws2.max_row)
        dict = []
        rows = ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=6)
        for c1, c2, c3, c4, c5, c6 in rows:
            d = {'id_contract': id_contract,'id_name': c1.value, 'id_product': c2.value, 'cost': round(c5.value, 2),
                 'amount': round(c4.value, 2), 'total': round(c6.value, 2), 'created': datetime.datetime.now(),
                 'update': datetime.datetime.now(), 'id_refs': id_refs}
            print(d)
            dict.append(d)
        mod_sql_item_orders(dict)

    def migration_service_postgres(self):
        wb = openpyxl.load_workbook("C:\\Users\\Master\\Desktop\\РАБОТА\\Transfer_to_Postgres\\trans_sql.xlsx")
        ws1 = wb['sett']
        id_contract = copy.copy(ws1['a2'].value)
        id_refs = copy.copy(ws1['b2'].value)
        ws2 = wb['service']
        print(ws2.max_row)
        d_service = []
        d_pay = []
        rows = ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=6)
        for c1, c2, c3, c4, c5, c6 in rows:
            d = {'id_contract': id_contract, 'service_date': c1.value, 'id_provider': c3.value,
                 'name_service': c4.value, 'cost_service': round(c5.value, 2),
                 'temp_col': c6.value, 'created': datetime.datetime.now(),
                 'update': datetime.datetime.now(), 'id_adress_refreshment': id_refs}
            d_service.append(d)
        print(d_service)
        mod_sql_service(d_service, flgDebug)
        ws_s_pay = wb['service_pay']
        print(ws2.max_row)
        rows_pay = ws_s_pay.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=5)
        for c1, c2, c3, c4, c5 in rows_pay:
            d = {'temp_col': c1.value, 'id_company': c2.value, 'invoice_date': c3.value,
                 'invoice_number': c4.value, 'invoice_cost': round(c5.value, 2),
                 'created': datetime.datetime.now(), 'update': datetime.datetime.now()}
            d_pay.append(d)
        print(d_pay)
        mod_sql_service_pay(d_pay, flgDebug)

    def migration_reiburs_postgres(self):
        wb = openpyxl.load_workbook("C:\\Users\\Master\\Desktop\\РАБОТА\\Transfer_to_Postgres\\trans_sql.xlsx")
        ws1 = wb['sett']
        d_pay = []
        id_contract = copy.copy(ws1['a2'].value)
        ws2 = wb['reimburs']
        print(ws2.max_row)
        rows = ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=4)
        for c1, c2, c3, c4 in rows:
            d = {'id_contract': id_contract, 'id_company': c1.value, 'name_reimburs': c2.value,
                 'cost_reimburs': round(c3.value, 2), 'date_reimburs': c4.value, 'created': datetime.datetime.now(),
                 'update': datetime.datetime.now()}
            d_pay.append(d)
        mod_sql_reimburs(d_pay, flgDebug)
        pass

    def migration_contractor_postgres(self):
        wb = openpyxl.load_workbook("C:\\Users\\Master\\Desktop\\РАБОТА\\Transfer_to_Postgres\\trans_sql.xlsx")
        ws1 = wb['sett']
        id_contract = copy.copy(ws1['a2'].value)
        id_refs = copy.copy(ws1['b2'].value)
        ws2 = wb['contractor']
        print(ws2.max_row)
        d_contract = []
        d_pay = []
        rows = ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=8)
        for c1, c2, c3, c4, c5, c6, c7, c8 in rows:
            d = {'id_contract': id_contract, 'id_provider': c4.value, 'id_customer': c8.value,
                 'name_contract': c5.value, 'contract_num': c1.value, 'contract_date': c2.value,
                 'cost_contract': round(c6.value, 2), 'created': datetime.datetime.now(),
                 'update': datetime.datetime.now(), 'status': 'not active', 'id_adress_refreshment': id_refs}
            d_contract.append(d)
        print(d_contract)
        mod_sql_contractor(d_contract, flgDebug)

        ws_s_pay = wb['contractor_pay']
        rows_pay = ws_s_pay.iter_rows(min_row=2, max_row=ws_s_pay.max_row, min_col=1, max_col=3)
        for c1, c2, c3 in rows_pay:
            d = {'id_contract': id_contract, 'contract_num': c3.value, 'date_payment': c1.value,
                 'cost_payment': c2.value, 'created': datetime.datetime.now(), 'update': datetime.datetime.now()}
            d_pay.append(d)
        print(d_pay)
        mod_sql_contractor_pay(d_pay, flgDebug)

    def migration_rent_postgres(self):
        wb = openpyxl.load_workbook("C:\\Users\\Master\\Desktop\\РАБОТА\\Transfer_to_Postgres\\trans_sql.xlsx")
        ws1 = wb['sett']
        id_contract = copy.copy(ws1['a2'].value)
        id_refs = copy.copy(ws1['b2'].value)
        ws2 = wb['rent']
        d_rent = []
        rows = ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=6)
        for c1, c2, c3, c4, c5, c6 in rows:
            d = {'id_contract': id_contract, 'id_adress_refreshment': id_refs, 'rent_date': c1.value,
                 'rent_name': c2.value,'rent_cost': c3.value, 'id_company': c4.value, 'rent_num': c5.value,
                 'id_provider': c6.value, 'created': datetime.datetime.now(), 'update': datetime.datetime.now()}
            d_rent.append(d)
        mod_sql_rent(d_rent)

flgDebug = False
# CoreXLSX.migration_order_in_postgres('w');
# CoreXLSX.migration_payment_in_postgres('w');
# CoreXLSX.migration_items_in_postgres('w');
# # # # # # #
# CoreXLSX.migration_service_postgres('w')
# # CoreXLSX.migration_reiburs_postgres('w')
# CoreXLSX.migration_contractor_postgres('w')
# CoreXLSX.migration_rent_postgres('w')


def check_petrovich(str_1):
    symbol = '"'
    print(str_1.value)
    if str_1.value:
        for r in str_1.value.split():
            if "Петрович" in r.replace(symbol, ""):
                return "ok"
        return "err"
    return "err"



# CoreXLSX.migration_product_in_postgres('w');







# wb.create_sheet("new_sheet")
# wb.save("Code_Product.xlsx")
# print(ws['b5'].value)
# print(ws.cell(row=6, column=3).value)
# value_range = ws['a2':'b12']
# print(value_range)
# for a,b in value_range:
#     print(a.value, b.value)
# for row in rows:
#     print(row)
